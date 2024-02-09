import BAC0
from BAC0.core.devices.local.models import (analog_input,analog_output,analog_value,binary_input,binary_output,binary_value,character_string,date_value,datetime_value,humidity_input,humidity_value,make_state_text,multistate_input,multistate_output,multistate_value,temperature_input,temperature_value,)
from BAC0.core.devices.local.object import ObjectFactory
import re
import time
import configparser
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np
import os


### Logging Settings ###
# Silence (use CRITICAL so not much messages will be sent) 
BAC0.log_level('silence') 

# Verbose 
# BAC0.log_level('info') 

# # Default, Info on console….but Warning in file 
# BAC0.log_level(log_file='info', stdout='info', stderr='critical')

# Debug in file and console… this is a bad idea as the console will be filled 
# BAC0.log_level(file='debug', stdout='debug', stderr='critical')


### CLASSES ###
class Device:
    def __init__(self, address, device_instance, ip_address, net, mac):
        self.address = address
        self.deviceInstance = device_instance
        self.ipAddress = ip_address
        self.net = net
        self.mac = mac


    def __repr__(self):
        return f"Device(deviceInstance={self.deviceInstance}, ipAddress={self.ipAddress}, net={self.net}, mac={self.mac})"


### BACnet FUNCTIONS ###
def bacnetInitialize():
    config = configparser.ConfigParser()
    config.read('settings.ini')
    ipAddress = config.get('bacnet', 'ipAddress')
    udpPort = config.get('bacnet', 'udpPort')
    bacnet = BAC0.lite(ip=ipAddress, port=udpPort)
    
    return bacnet

def deviceScan(bacnet):
    config = configparser.ConfigParser()
    config.read('settings.ini')
    device_ranges = config.get('bacnet', 'deviceRanges').split(';')
    scanTimeout = config.getint('bacnet', 'scanTimeout')

    deviceList = []

    for device_range in device_ranges:
        range_limits = [int(limit) for limit in device_range.split('-') if limit.isdigit()]

        if len(range_limits) == 1:
            startInstance = endInstance = range_limits[0]
        else: 
            startInstance, endInstance = range_limits
        passes = 1

        noNewDeviceCounter = 0
        while noNewDeviceCounter < scanTimeout:
            print(f"Scan for devices {startInstance} to {endInstance}, Pass # {passes}")
            newDeviceFlag = False
            bacnet.discover(limits=(startInstance, endInstance))
            deviceDict = bacnet.discoveredDevices

            for key, value in deviceDict.items():
                address = key[0]
                deviceInstance = key[1]
                ipAddress = ""
                net = ""
                mac = ""

                if re.match(r'^(\d{1,3}\.){3}\d{1,3}$', address):
                    ipAddress = address
                else:
                    address_parts = address.split(':')
                    if len(address_parts) == 2:
                        net = address_parts[0]
                        mac = address_parts[1]

                device = Device(address, deviceInstance, ipAddress, net, mac)
                if all(device.deviceInstance != d.deviceInstance for d in deviceList):
                    deviceList.append(device)
                    print("*** added " + str(device.deviceInstance))
                    newDeviceFlag = True

            if newDeviceFlag:
                noNewDeviceCounter = 0
            else:
                noNewDeviceCounter += 1

            passes += 1

    deviceList = sorted(deviceList, key=lambda x: x.deviceInstance)

    # Create a dictionary to store device information
    data = {
        'address': [device.address for device in deviceList],
        'deviceInstance': [device.deviceInstance for device in deviceList],
        'IP': [device.ipAddress for device in deviceList],
        'Network': [device.net for device in deviceList],
        'MAC': [device.mac for device in deviceList]
    }

    # Create a Pandas DataFrame from the device information
    df = pd.DataFrame(data)

    # Write the DataFrame to an Excel file
    df.to_excel('device_info.xlsx', index=False)

    print("Device scan complete.")

def readAv(bacnet):
    # Read the existing Device Excel into a DataFrame
    df = pd.read_excel('device_info.xlsx')

    # Create a new DataFrame to store AV values
    av_df = pd.DataFrame(columns=['deviceInstance'])

    # Get unique device instances
    device_instances = df['deviceInstance'].unique()
    av_df['deviceInstance'] = device_instances

    # Iterate through AV ranges from the .ini file
    config = configparser.ConfigParser()
    config.read('settings.ini')
    avRangeList = config.get('bacnet', 'avRange').split(';')

    for avRange in avRangeList:
        # Check if it's a single AV or a range
        if '-' in avRange:
            start, end = avRange.split('-')
            startAV, endAV = int(start), int(end)
        else:
            startAV = endAV = int(avRange)

        for av_number in range(startAV, endAV + 1):
            av_values = []
            for device_instance in device_instances:
                address = df[df['deviceInstance'] == device_instance]['address'].values[0]

                try:
                    av_value = bacnet.read(f'{address} analogValue {av_number} presentValue')
                    # Round the AV value to 3 decimal points
                    av_value_rounded = round(av_value, 3) if av_value is not None else None
                    av_values.append(av_value_rounded)
                except Exception as e:
                    print(f"Error reading AV{av_number} for device {address}: {e}")
                    av_values.append(None)

            # Add AV values as new columns in the DataFrame
            av_df[f'AV{av_number}'] = av_values

    # Write AV values to a new Excel file
    av_df.to_excel('av_values.xlsx', index=False)

    highlight_outliers()

    return av_df

def readBv(bacnet):
    # Read the existing Device Excel into a DataFrame
    df = pd.read_excel('device_info.xlsx')

    # Create a new DataFrame to store BV values
    bv_df = pd.DataFrame(columns=['deviceInstance'])

    # Get unique device instances
    device_instances = df['deviceInstance'].unique()
    bv_df['deviceInstance'] = device_instances

    # Iterate through BV ranges from the .ini file
    config = configparser.ConfigParser()
    config.read('settings.ini')
    bvRange = config.get('bacnet', 'bvRange').split(';')

    for bvRange in bvRange:
        # Check if it's a single AV or a range
        if '-' in bvRange:
            start, end = bvRange.split('-')
            startBV, endBV = int(start), int(end)
        else:
            startBV = endBV = int(bvRange)


        for bv_number in range(startBV, endBV + 1):
            bv_values = []
            for device_instance in device_instances:
                address = df[df['deviceInstance'] == device_instance]['address'].values[0]

                try:
                    bv_value = bacnet.read(f'{address} binaryValue {bv_number} presentValue')
                    bv_values.append(bv_value)
                except Exception as e:
                    print(f"Error reading BV{bv_number} for device {address}: {e}")
                    bv_values.append(None)

            # Add BV values as new columns in the DataFrame
            bv_df[f'BV{bv_number}'] = bv_values

    # Write BV values to a new Excel file
    bv_df.to_excel('bv_values.xlsx', index=False)

    return bv_df



### Other helper functions ###
def highlight_outliers():
    file_path = os.path.join(os.getcwd(), 'av_values.xlsx')

    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Get the data from the sheet
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Convert data to NumPy array
    data = np.array(data[1:], dtype=np.float64)  # Exclude header row

    # Calculate standard deviation for each column
    std_dev = np.nanstd(data, axis=0)
    mean = np.nanmean(data, axis=0)

    # Highlight cells that are more than 2 standard deviations away in red
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for col_index, (column, std, avg) in enumerate(zip(data.T, std_dev, mean), start=1):
        for cell_index, value in enumerate(column, start=2):
            cell = sheet.cell(row=cell_index, column=col_index)
            if value > avg + 1 * std or value < avg - 1 * std:
                cell.fill = yellow_fill
            if value > avg + 2 * std or value < avg - 2 * std:
                cell.fill = red_fill

    # Save the updated workbook
    wb.save(file_path)


def main():
    bacnet = bacnetInitialize()
    # deviceScan(bacnet)
    readAv(bacnet)
    # readBv(bacnet)

    return

if __name__ == '__main__':
    main()
