import BAC0
from BAC0.core.devices.local.models import (
    analog_input,
    analog_output,
    analog_value,
    binary_input,
    binary_output,
    binary_value,
    character_string,
    date_value,
    datetime_value,
    humidity_input,
    humidity_value,
    make_state_text,
    multistate_input,
    multistate_output,
    multistate_value,
    temperature_input,
    temperature_value,
)
from BAC0.core.devices.local.object import ObjectFactory
import re
import time
import configparser
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np
import os
import re
import warnings
import logging
import math
import datetime
from dotenv import load_dotenv

### Logging Settings ###
# Silence (use CRITICAL so not much messages will be sent)
BAC0.log_level("silence")

### Ignore openpyxl data validation warning
warnings.simplefilter(action="ignore", category=UserWarning)

# Configure logging to write to a file named 'log.txt'
logging.basicConfig(
    filename="log.txt",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)

# Customer logging for just BACnet write operations
bacnet_logger = logging.getLogger("custom_logger")
bacnet_logger.setLevel(logging.INFO)
file_handler = logging.FileHandler("bacnet log.txt")
formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
file_handler.setFormatter(formatter)
bacnet_logger.addHandler(file_handler)


### CLASSES ###
class Device:
    def __init__(self, address, device_instance, ip_address, net, mac):
        self.address = address
        self.deviceInstance = device_instance
        self.ipAddress = ip_address
        self.net = net
        self.mac = mac

    def __repr__(self):
        return f"Device(deviceInstance={self.deviceInstance}, ipAddress={self.ipAddress}, net={self.net}, mac={self.mac})"


### FUNCTIONS ###
def authenticate():
    load_dotenv()

    # Define the correct password
    correct_password = os.getenv("PASSWORD")

    # Prompt the user for a password
    password = input("Enter the password to run the program: ")

    # Check if the entered password matches the correct password
    if password != correct_password:
        print("Incorrect password. Access denied.")
        return False

    # Check date / time
    current_date = datetime.datetime.now()
    kill_date = datetime.datetime(2024, 4, 1)

    if current_date > kill_date:
        return False

    return True


def bacnet_initialize():
    """
    Parameters: None
    Takes in BACnet configuration parameters from settings.ini
    Returns: BACnet device

    REV History:
    2024-02-08 (mikes): initial
    """
    # Takes in BACnet configuration parameters from settings.ini
    # Creates and returns a BACnet device
    config = configparser.ConfigParser()
    config.read("settings.ini")
    ipAddress = config.get("bacnet", "ipAddress")
    udpPort = config.get("bacnet", "udpPort")
    bacnet = BAC0.lite(ip=ipAddress, port=udpPort)
    return bacnet


def build_device_manager(bacnet, DI_list):
    """
    Parameters: bacnet device, list of Device Instances
    Calls device_scan() once for entire range, then scans for missing devices
    Return: Pandas df with address information for each Device Instance

    REV History:
    2024-02-08 (mikes): initial
    """

    # Scan for entire range from max to min
    device_manager = device_scan(bacnet, min(DI_list), max(DI_list))
    # device_manager = device_scan(bacnet, min(DI_list), 1200)

    # Scan for missing items
    for device in DI_list:
        if device not in np.array(device_manager["deviceInstance"].values, dtype=np.int64):
            print(f"Scanning for missing device {device}")
            device_info = device_scan(bacnet, device, device)
            if device_info is not None:  # Check if device_info is not empty
                device_manager = pd.concat([device_manager, device_info], ignore_index=True)

    # Remove duplicate rows based on 'deviceInstance' column
    device_manager = device_manager.drop_duplicates(subset="deviceInstance", keep="first").reset_index(drop=True)

    # Remove deviees that aren't in DI_list
    device_manager = device_manager[device_manager["deviceInstance"].isin(DI_list)]

    return device_manager


def device_scan(bacnet, start_instance, end_instance):
    """
    Parameters: bacnet device, list of Device Instances
    Conducts a scan for a set range of device instances.  Will repeat scans based on settings.ini
    Return: Pandas df with address information for each Device Instance

    REV History:
    2024-02-08 (mikes): initial
    """

    config = configparser.ConfigParser()
    config.read("settings.ini")
    scanTimeout = config.getint("bacnet", "scanTimeout")

    device_manager = []
    passes = 1

    while passes <= scanTimeout:
        # If single instance scan and device is found, break
        if start_instance == end_instance and any(device.deviceInstance == start_instance for device in device_manager):
            break

        print(f"Scan for devices {start_instance} to {end_instance}, Pass # {passes}")

        # Scan for device
        bacnet.discover(limits=(start_instance, end_instance))
        device_dict = bacnet.discoveredDevices

        for key, value in device_dict.items():
            address = key[0]
            deviceInstance = key[1]
            ipAddress = ""
            net = ""
            mac = ""

            if re.match(r"^(\d{1,3}\.){3}\d{1,3}$", address):
                ipAddress = address
            else:
                address_parts = address.split(":")
                if len(address_parts) == 2:
                    net = address_parts[0]
                    mac = address_parts[1]

            device = Device(address, deviceInstance, ipAddress, net, mac)
            if all(device.deviceInstance != d.deviceInstance for d in device_manager):
                device_manager.append(device)
                print("** Found device " + str(device.deviceInstance))
                device_found = True

        passes += 1

    device_manager = sorted(device_manager, key=lambda x: x.deviceInstance)

    # Create a dictionary to store device information
    data = {
        "address": [device.address for device in device_manager],
        "deviceInstance": [device.deviceInstance for device in device_manager],
        "IP": [device.ipAddress for device in device_manager],
        "Network": [device.net for device in device_manager],
        "MAC": [device.mac for device in device_manager],
    }

    # Create a Pandas DataFrame from the device information
    df = pd.DataFrame(data)

    return df


def scan_device_objects(bacnet):
    return


def read_point(bacnet, device_manager, device_instance, object_type, object_instance, property, index=None):
    """
    Parameters: lots
    Performs BACnet read
    Return: BACnet value.  If error, returns "NR"

    Example:
    value = read_point(bacnet, device_manager, 1001, "binaryOutput", "0", "priorityArray", 14)

    REV History:
    2024-02-08 (mikes): initial
    """

    # Variables
    address = None
    value = None

    # Get bacnet address from device_manager
    if device_instance in device_manager["deviceInstance"].values:
        # Get the address for device_instance 1000
        address = device_manager.loc[device_manager["deviceInstance"] == device_instance, "address"].iloc[0]
    else:
        return "NR"

    # Read BACnet point
    try:
        # Read DDC file name
        if object_type == "program":
            value = bacnet.read(f"{address} program 0 {property}")

        # Read from device
        elif object_type == "device":
            value = bacnet.read(f"{address} device {device_instance} {property}")

        # Read point
        else:
            value = bacnet.read(f"{address} {object_type} {object_instance} {property}")

    except Exception as e:
        logging.error(
            f"read_point error.  error: {e} device: {device_instance} object_type: {object_type} object_instance: {object_instance} property: {property} index: {index}"
        )
        value = "NR"
        return value

    # Unpack priority array
    if property == "priorityArray":
        array = serialize_priority_array(value.dict_contents(), object_type)
        if str(index) in array.keys():
            value = array[str(index)]
        else:
            value = "NR"

    return value


def output_to_excel(df, DI_list, points_list, sheet_name):
    """
    Parameters: pandas df, other sorting information
    Writes data from df to excel.
    Return: None

    REV History:
    2024-02-08 (mikes): initial
    """
    # Take in dataframe read from BACnet
    # Write to excel

    file_name = "Point Read Write.xlsx"

    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_name)

        # Select the active sheet
        sheet = wb[sheet_name]

        for device_instance in DI_list:
            # Find df row index
            if sheet_name == "read":
                df_row_index = df.index[df["READ"] == device_instance].tolist()[0]
            else:
                df_row_index = df.index[df["WRITE"] == device_instance].tolist()[0]

            # Find Excel row index
            excel_row_index = None
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
                if row[0].value == device_instance:
                    excel_row_index = row[0].row
                    break

            # Iterate through columns
            for point in points_list:
                col_letter = point["col_letter"]
                col_name = point["col_index"]

                df_value = df.at[df_row_index, col_name]
                cell = str(col_letter) + str(excel_row_index)
                sheet[cell] = df_value

        # Save the workbook
        wb.save(file_name)
        logging.info(f"output_to_excel successful")

    except Exception as e:
        logging.error(f"output_to_excel error: {e}")


def main():
    # Authenticate the user
    # if not authenticate():
    #     return
    bacnet = BAC0.connect("192.168.105.40/24")

    device_instance = 1001

    # Initialize BACnet
    # bacnet = bacnet_initialize()
    bacnet.whois('1000 1001')


    # Device Scan
    # device_manager = device_scan(bacnet, 1000, 1001)
    print(bacnet.devices)

    # Define / register a controller
    myController = BAC0.device('192.168.105.7', 1000, bacnet)
    print(myController.points)

    return

    # Prompt for device instance.  Do this in tkinter
    device_instance = 1001

    # Device scan for that device instance
    device_manager = device_scan(bacnet, device_instance, device_instance)
    print(device_manager)

    print(bacnet.bacnet_properties())

    # Get list of points supported

    # BACnet read: object name, object description, present value

    # Output to excel


if __name__ == "__main__":
    main()
