import BAC0
from BAC0.core.devices.local.models import (
    analog_input,
    analog_output,
    analog_value,
    binary_input,
    binary_output,
    binary_value,
    character_string,
    date_value,
    datetime_value,
    humidity_input,
    humidity_value,
    make_state_text,
    multistate_input,
    multistate_output,
    multistate_value,
    temperature_input,
    temperature_value,
)
from BAC0.core.devices.local.object import ObjectFactory
import re
import time
import configparser
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np
import os
import re
import warnings
import logging
import math
import datetime


### Logging Settings ###
# Silence (use CRITICAL so not much messages will be sent)
BAC0.log_level("silence")

### Ignore openpyxl data validation warning
warnings.simplefilter(action="ignore", category=UserWarning)

# Configure logging to write to a file named 'log.txt'
logging.basicConfig(
    filename="log.txt",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)

# Customer logging for just BACnet write operations
bacnet_logger = logging.getLogger("custom_logger")
bacnet_logger.setLevel(logging.INFO)
file_handler = logging.FileHandler("bacnet log.txt")
formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
file_handler.setFormatter(formatter)
bacnet_logger.addHandler(file_handler)


### CLASSES ###
class Device:
    def __init__(self, address, device_instance, ip_address, net, mac):
        self.address = address
        self.deviceInstance = device_instance
        self.ipAddress = ip_address
        self.net = net
        self.mac = mac

    def __repr__(self):
        return f"Device(deviceInstance={self.deviceInstance}, ipAddress={self.ipAddress}, net={self.net}, mac={self.mac})"


### FUNCTIONS ###
def authenticate():
    # Define the correct password
    correct_password = "SithLordShen"

    # Prompt the user for a password
    password = input("Enter the password to run the program: ")

    # Check if the entered password matches the correct password
    if password == correct_password:
        return True
    else:
        print("Incorrect password. Access denied.")
        return False

def bacnet_initialize():
    """
    Parameters: None
    Takes in BACnet configuration parameters from settings.ini
    Returns: BACnet device

    REV History:
    2024-02-08 (mikes): initial
    """
    # Takes in BACnet configuration parameters from settings.ini
    # Creates and returns a BACnet device
    config = configparser.ConfigParser()
    config.read("settings.ini")
    ipAddress = config.get("bacnet", "ipAddress")
    udpPort = config.get("bacnet", "udpPort")
    bacnet = BAC0.lite(ip=ipAddress, port=udpPort)
    return bacnet


def read_from_excel():
    """
    Paramters: None
    Returns: pandas dataframe as read from excel file

    REV History:
    2024-02-08 (mikes): initial
    """

    file_name = "Point Read Write.xlsx"

    try:
        # Read the Excel file into a pandas DataFrame
        read_df = pd.read_excel(file_name, sheet_name="read")
        write_df = pd.read_excel(file_name, sheet_name="write")

    except Exception as e:
        return f"Error reading Excel file: {e}"

    return read_df, write_df


def get_di_list(excel_df):
    """
    Parameters: dataframe that's read from excel
    Return: List of device instances

    REV History:
    2024-02-08 (mikes): initial
    """

    DI_list = []
    for i in range(2, len(excel_df)):
        cell_value = excel_df.iloc[i, 0]
        # Check if cell value is numeric
        if pd.notnull(cell_value) and isinstance(cell_value, (int, float)):
            DI_list.append(cell_value)
    return DI_list


def get_points_list(df):
    """
    Parameters: dataframe that's read from excel
    Return: List of dicts.  Each dict contains point information for reading / writing

    REV History:
    2024-02-08 (mikes): initial
    """

    # Load the workbook
    file_name = "Point Read Write.xlsx"
    try:
        wb = openpyxl.load_workbook(file_name)
    except Exception as e:
        logging.error(f"Error reading Excel file: {e}")
        return

    if "READ" in df.columns:
        df.rename(columns={"READ": "A1"}, inplace=True)
        sheet = wb["read"]
    else:
        df.rename(columns={"WRITE": "A1"}, inplace=True)
        sheet = wb["write"]

    # Get row indexes
    object_type_row = df[df["A1"] == "object_type"].index[0]
    object_instance_row = df[df["A1"] == "object_instance"].index[0]
    object_property_row = df[df["A1"] == "object_property"].index[0]
    index_row = df[df["A1"] == "index"].index[0]

    # Get column list
    column_list = df.columns.tolist()
    column_list.remove("A1")

    points_list = []

    # Iterate through columns and create points list
    for column_name in column_list:
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == column_name:
                column_letter = openpyxl.utils.get_column_letter(col)
                break

        point = {
            "col_index": column_name,
            "object_type": df.at[object_type_row, column_name],
            "object_instance": df.at[object_instance_row, column_name],
            "property": df.at[object_property_row, column_name],
            "index": df.at[index_row, column_name],
            "col_letter": column_letter,
        }

        points_list.append(point)

    wb.close()

    return points_list


def build_device_manager(bacnet, DI_list):
    """
    Parameters: bacnet device, list of Device Instances
    Calls device_scan() once for entire range, then scans for missing devices
    Return: Pandas df with address information for each Device Instance

    REV History:
    2024-02-08 (mikes): initial
    """

    # Scan for entire range from max to min
    device_manager = device_scan(bacnet, min(DI_list), max(DI_list))
    # device_manager = device_scan(bacnet, min(DI_list), 1200)

    # Scan for missing items
    for device in DI_list:
        if device not in np.array(device_manager["deviceInstance"].values, dtype=np.int64):
            print(f"Scanning for missing device {device}")
            device_info = device_scan(bacnet, device, device)
            if device_info is not None:  # Check if device_info is not empty
                device_manager = pd.concat([device_manager, device_info], ignore_index=True)

    # Remove duplicate rows based on 'deviceInstance' column
    device_manager = device_manager.drop_duplicates(subset="deviceInstance", keep="first").reset_index(drop=True)

    # Remove deviees that aren't in DI_list
    device_manager = device_manager[device_manager["deviceInstance"].isin(DI_list)]

    return device_manager


def device_scan(bacnet, start_instance, end_instance):
    """
    Parameters: bacnet device, list of Device Instances
    Conducts a scan for a set range of device instances.  Will repeat scans based on settings.ini
    Return: Pandas df with address information for each Device Instance

    REV History:
    2024-02-08 (mikes): initial
    """

    config = configparser.ConfigParser()
    config.read("settings.ini")
    scanTimeout = config.getint("bacnet", "scanTimeout")

    device_manager = []
    passes = 1

    while passes <= scanTimeout:
        # If single instance scan and device is found, break
        if start_instance == end_instance and any(device.deviceInstance == start_instance for device in device_manager):
            break

        print(f"Scan for devices {start_instance} to {end_instance}, Pass # {passes}")

        # Scan for device
        bacnet.discover(limits=(start_instance, end_instance))
        device_dict = bacnet.discoveredDevices

        for key, value in device_dict.items():
            address = key[0]
            deviceInstance = key[1]
            ipAddress = ""
            net = ""
            mac = ""

            if re.match(r"^(\d{1,3}\.){3}\d{1,3}$", address):
                ipAddress = address
            else:
                address_parts = address.split(":")
                if len(address_parts) == 2:
                    net = address_parts[0]
                    mac = address_parts[1]

            device = Device(address, deviceInstance, ipAddress, net, mac)
            if all(device.deviceInstance != d.deviceInstance for d in device_manager):
                device_manager.append(device)
                print("** Found device " + str(device.deviceInstance))
                device_found = True

        passes += 1

    device_manager = sorted(device_manager, key=lambda x: x.deviceInstance)

    # Create a dictionary to store device information
    data = {
        "address": [device.address for device in device_manager],
        "deviceInstance": [device.deviceInstance for device in device_manager],
        "IP": [device.ipAddress for device in device_manager],
        "Network": [device.net for device in device_manager],
        "MAC": [device.mac for device in device_manager],
    }

    # Create a Pandas DataFrame from the device information
    df = pd.DataFrame(data)

    return df


def serialize_priority_array(priority_array, object_type):
    """
    Parameters: priority_array object, object_type
    Unpacks priority_array object.
    Replaces binary text with active and inactive
    Return: dict with priority array data

    REV History:
    2024-02-08 (mikes): initial
    """

    priority_array_dict = {}

    for i in range(16):
        sub_dict = priority_array[i]
        if sub_dict:  # Check if sub-dictionary is not empty
            value = list(sub_dict.values())[0]

            # Replace null text
            if (not value) and (value != 0):
                value = "null"

            # Replace binary text with "active" and "inactive"
            if object_type == "binaryInput" or object_type == "binaryOutput" or object_type == "binaryValue":

                if value == 1:
                    value = "active"
                elif value == 0:
                    value = "inactive"

            priority_array_dict[str(i + 1)] = value
        else:
            priority_array_dict[str(i + 1)] = "null"
    return priority_array_dict


def read_point(bacnet, device_manager, device_instance, object_type, object_instance, property, index=None):
    """
    Parameters: lots
    Performs BACnet read
    Return: BACnet value.  If error, returns "NR"

    Example:
    value = read_point(bacnet, device_manager, 1001, "binaryOutput", "0", "priorityArray", 14)
    
    REV History:
    2024-02-08 (mikes): initial
    """

    # Variables
    address = None
    value = None

    # Get bacnet address from device_manager
    if device_instance in device_manager["deviceInstance"].values:
        # Get the address for device_instance 1000
        address = device_manager.loc[device_manager["deviceInstance"] == device_instance, "address"].iloc[0]
    else:
        return "NR"

    # Read BACnet point
    try:
        # Read DDC file name
        if object_type == "program":
            value = bacnet.read(f"{address} program 0 {property}")

        # Read from device
        elif object_type == "device":
            value = bacnet.read(f"{address} device {device_instance} {property}")

        # Read point
        else:
            value = bacnet.read(f"{address} {object_type} {object_instance} {property}")

    except Exception as e:
        logging.error(
            f"read_point error.  error: {e} device: {device_instance} object_type: {object_type} object_instance: {object_instance} property: {property} index: {index}"
        )
        value = "NR"
        return value

    # Unpack priority array
    if property == "priorityArray":
        array = serialize_priority_array(value.dict_contents(), object_type)
        if str(index) in array.keys():
            value = array[str(index)]
        else:
            value = "NR"

    return value


def write_point(bacnet, device_manager, device_instance, object_type, object_instance, property, value, index=None):
    """
    Parameters: lots
    Performs BACnet write
    Logs previous value prior to writing
    Return: None

    Example:
    write_point(bacnet, device_manager, 1001, "binaryOutput", "20", "priorityArray", "inactive", 3)
    
    REV History:
    2024-02-08 (mikes): initial
    """

    # Variables
    address = None
    value = re.sub(r"\s+", "_", str(value))

    # Get bacnet address from device_manager
    if device_instance in device_manager["deviceInstance"].values:
        # Get the address for device_instance 1000
        address = device_manager.loc[device_manager["deviceInstance"] == device_instance, "address"].iloc[0]
    else:
        return

    # Don't allow writing to program
    if object_type == "program":
        return

    # Read BACnet point and log value
    read_value = read_point(bacnet, device_manager, device_instance, object_type, object_instance, property, index)

    if index is None:
        bacnet_logger.info(f"Writing to {device_instance}:{object_type}{object_instance} {property}.  Original value: {read_value}")
    else:
        bacnet_logger.info(
            f"Writing to {device_instance}:{object_type}{object_instance} {property} priority {index}.  Original value: {read_value}"
        )

    # Write BACnet point
    try:
        # Write to device
        if object_type == "device":
            bacnet.write(f"{address} device {device_instance} {property} {value}")

        # Write to priority array
        elif property == "priorityArray":
            bacnet.write(f"{address} {object_type} {object_instance} presentValue {value} - {index}")

        # Write to point
        else:
            bacnet.write(f"{address} {object_type} {object_instance} {property} {value}")

    except Exception as e:
        logging.error(
            f"write_point error.  error: {e} device: {device_instance} object_type: {object_type} object_instance: {object_instance} property: {property} index: {index}"
        )
        bacnet_logger.error(
            f"write_point error.  error: {e} device: {device_instance} object_type: {object_type} object_instance: {object_instance} property: {property} index: {index}"
        )

        return None
    return


def output_to_excel(df, DI_list, points_list, sheet_name):
    """
    Parameters: pandas df, other sorting information
    Writes data from df to excel.
    Return: None

    REV History:
    2024-02-08 (mikes): initial
    """
    # Take in dataframe read from BACnet
    # Write to excel

    file_name = "Point Read Write.xlsx"

    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_name)

        # Select the active sheet
        sheet = wb[sheet_name]

        for device_instance in DI_list:
            # Find df row index
            if sheet_name == "read":
                df_row_index = df.index[df["READ"] == device_instance].tolist()[0]
            else:
                df_row_index = df.index[df["WRITE"] == device_instance].tolist()[0]

            # Find Excel row index
            excel_row_index = None
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
                if row[0].value == device_instance:
                    excel_row_index = row[0].row
                    break

            # Iterate through columns
            for point in points_list:
                col_letter = point["col_letter"]
                col_name = point["col_index"]

                df_value = df.at[df_row_index, col_name]
                cell = str(col_letter) + str(excel_row_index)
                sheet[cell] = df_value

        # Save the workbook
        wb.save(file_name)
        logging.info(f"output_to_excel successful")

    except Exception as e:
        logging.error(f"output_to_excel error: {e}")


def execute_read():
    """
    Parameters: none
    Main call to read parameters from excel and reads BACnet data
    Return: writes data back to same excel file

    REV History:
    2024-02-08 (mikes): initial
    """

    bacnet = bacnet_initialize()
    read_df, write_df = read_from_excel()
    DI_list = get_di_list(read_df)
    device_manager = build_device_manager(bacnet, DI_list)

    # Make df for read / write similar
    df = read_df
    points_list = get_points_list(df)
    df.rename(columns={"READ": "A1"}, inplace=True)

    for device_instance in DI_list:
        print(f"Reading from {device_instance}...")
        # find row index for this device instance
        row_index = df.index[df["A1"] == device_instance].tolist()[0]

        # Iterate through columns
        for point in points_list:
            # Read the BACnet point
            value = read_point(
                bacnet, device_manager, device_instance, point["object_type"], point["object_instance"], point["property"], point["index"]
            )

            # Write to df
            col_index = point["col_index"]
            df.at[row_index, col_index] = value

    df.rename(columns={"A1": "READ"}, inplace=True)

    # Write back to excel sheet "read"
    output_to_excel(df, DI_list, points_list, "read")

    return


def execute_write():
    """
    Parameters: none
    Main call to write BACnet parameters to excel
    Return: none

    REV History:
    2024-02-08 (mikes): initial
    """

    bacnet = bacnet_initialize()
    read_df, write_df = read_from_excel()

    DI_list = get_di_list(write_df)
    device_manager = build_device_manager(bacnet, DI_list)

    # Make df for read / write similar
    df = write_df
    points_list = get_points_list(df)
    df.rename(columns={"WRITE": "A1"}, inplace=True)

    # Iterate through rows
    for device_instance in DI_list:
        print(f"Writing to {device_instance}...")

        # Find row index
        df_row_index = df.index[df["A1"] == device_instance].tolist()[0]

        # Iterate through points
        for point in points_list:
            # Find column name
            col_name = point["col_index"]

            # Find value to write from df
            df_value = df.at[df_row_index, col_name]
            if df_value == "auto":
                df_value = "null"

            if not (isinstance(df_value, float) and math.isnan(df_value)):
                # Write to BACnet
                print("here")
                write_point(
                    bacnet,
                    device_manager,
                    device_instance,
                    point["object_type"],
                    point["object_instance"],
                    point["property"],
                    df_value,
                    point["index"],
                )

    return


def main():
    # Authenticate the user
    if not authenticate():
        return

    current_date = datetime.datetime.now()
    kill_date = datetime.datetime(2024, 3, 1)

    if current_date > kill_date:
        print("Dead!")
        return

    execute_read()
    # execute_write()


if __name__ == "__main__":
    main()
